"""
Bhati March 2026 Phase 2 — Fetch & Analyse
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Architecture:
  - scans_cache.xlsx  = persistent DB for all scans (never lost on restart)
  - user_cache.xlsx   = persistent DB for Firebase users (incremental)
  - Sewadar Details.xlsx = badge master (read-only)
  - Cold start: load Excel → 0 Firestore reads → server ready instantly
  - Every 30s: only fetch docs newer than last_fetch timestamp
  - Analysis runs only when new scans arrive

Logic fixes vs Phase 1:
  - PRESENT = has ANY IN scan (even without OUT) ✅
  - ABSENT = all registered sewadars minus those with IN today ✅
  - duplicate_flags and multi_scan_flags are separate, no overlap ✅
  - Re-entry detects any IN→OUT→IN cycle regardless of duplicates ✅
  - centre_expected uses centre field (not department field) ✅
  - duty_hours = last_out - first_in per sewadar per day ✅
  - dept completion % computed server-side ✅
  - arrival_slots: 30-min bucket chart data ✅
"""

import firebase_admin
from firebase_admin import credentials, firestore
import pandas as pd
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
import json, os, math, tempfile

# ═══════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════
SCAN_COLLECTION  = "Bhati-March-2026-Ph2"
USER_COLLECTION  = "User"
TIMESTAMP_FIELD  = "timestamp"
USER_CREATED_FIELD = "createdAt"
IST_OFFSET       = timedelta(hours=5, minutes=30)

_BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
def _p(f):        return os.path.join(_BASE_DIR, f)

SCANS_EXCEL      = os.environ.get("SCANS_EXCEL_PATH", "/data/scans_cache.xlsx")
USERS_EXCEL      = os.environ.get("USERS_EXCEL_PATH", "/data/user_cache.xlsx")
OUTPUT_JSON      = _p("data.json")
SEWADAR_EXCEL    = os.environ.get("SEWADAR_EXCEL_PATH", _p("Sewadar Details.xlsx"))

# ═══════════════════════════════════════════════════════════════
# FIREBASE INIT — once at import time
# ═══════════════════════════════════════════════════════════════
if not firebase_admin._apps:
    secret = os.environ.get("FIREBASE_SERVICE_ACCOUNT")
    if secret:
        sa = json.loads(secret)
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump(sa, f)
            tmp_path = f.name
        cred = credentials.Certificate(tmp_path)
    else:
        cred = credentials.Certificate(_p("service_account.json"))
    firebase_admin.initialize_app(cred)

db = firestore.client()

# ═══════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ═══════════════════════════════════════════════════════════════
SCAN_COLS = [
    "scan_id","badge_no","sewadar_name","sewadar_centre","satsang_point",
    "department","gender","type","timestamp","date","time",
    "scanned_by_badge","scanned_by_name","scanned_by_centre"
]
USER_COLS = ["uid","badge_no","display_name","created_at"]

def ensure_excel(path, sheets):
    """Create Excel file with named sheets if it doesn't exist."""
    if not os.path.exists(path):
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, headers in sheets.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        # meta sheet always gets a timestamp row placeholder
        wb.save(path)
        print(f"Created: {path}")

def read_sheet_as_df(path, sheet):
    """Read an Excel sheet into a DataFrame. Returns empty DF if sheet missing."""
    try:
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
        return df.fillna("")
    except Exception:
        return pd.DataFrame()

def read_meta(path):
    """Read meta sheet → dict. Returns {} if missing."""
    try:
        df = pd.read_excel(path, sheet_name="meta", dtype=str).fillna("")
        return dict(zip(df["key"], df["value"])) if "key" in df.columns else {}
    except Exception:
        return {}

def write_meta(path, meta_dict):
    """Overwrite meta sheet in existing Excel with key/value pairs."""
    try:
        wb = load_workbook(path)
        if "meta" in wb.sheetnames:
            del wb["meta"]
        ws = wb.create_sheet("meta")
        ws.append(["key", "value"])
        for k, v in meta_dict.items():
            ws.append([k, str(v)])
        wb.save(path)
    except Exception as e:
        print(f"[WARN] write_meta failed: {e}")

def append_rows_to_sheet(path, sheet_name, rows, cols):
    """Append new rows to an existing Excel sheet."""
    if not rows:
        return
    try:
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(cols)
        else:
            ws = wb[sheet_name]
        for row in rows:
            ws.append([row.get(c, "") for c in cols])
        wb.save(path)
    except Exception as e:
        print(f"[WARN] append_rows failed for {sheet_name}: {e}")

# ═══════════════════════════════════════════════════════════════
# BADGE LOOKUP — loaded ONCE at module level
# ═══════════════════════════════════════════════════════════════
badge_lookup      = {}
department_expected = {}
centre_expected   = {}

if os.path.exists(SEWADAR_EXCEL):
    try:
        df_sew = pd.read_excel(SEWADAR_EXCEL)
        df_sew["Badge Number"] = df_sew["Badge Number"].astype(str).str.strip()
        for _, row in df_sew.iterrows():
            b = row["Badge Number"]
            badge_lookup[b] = {
                "name":          str(row.get("Name of Sewadar", "")).strip(),
                "gender":        str(row.get("Gender", "")).strip(),
                "satsang_point": str(row.get("Satsang Point", "")).strip(),
                "centre":        str(row.get("Centre", "")).strip(),
                "department":    str(row.get("Deployed Department", "")).strip(),
            }
        print(f"Badge lookup loaded: {len(badge_lookup)}")
        # Build expected counts by department AND by centre
        for info in badge_lookup.values():
            dept = info.get("department","").strip()
            cent = info.get("centre","").strip()
            if dept: department_expected[dept] = department_expected.get(dept,0) + 1
            if cent: centre_expected[cent]      = centre_expected.get(cent,0) + 1
    except Exception as e:
        print(f"[WARN] Excel load error: {e}")
else:
    print(f"[WARN] Excel not found: {SEWADAR_EXCEL}")

def lookup_badge(badge_no):
    return badge_lookup.get(str(badge_no).strip(), {
        "name":"Unknown","gender":"","satsang_point":"","centre":"","department":""
    })

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
def to_ist_naive(ts):
    if hasattr(ts, "tzinfo") and ts.tzinfo:
        return ts.astimezone(timezone(IST_OFFSET)).replace(tzinfo=None)
    return ts

def clean(obj):
    if isinstance(obj, float) and math.isnan(obj): return ""
    if isinstance(obj, dict):  return {k: clean(v) for k, v in obj.items()}
    if isinstance(obj, list):  return [clean(i) for i in obj]
    return obj

def parse_ts(ts_str):
    """Parse ISO timestamp string → datetime (UTC-aware). Returns None on failure."""
    if not ts_str:
        return None
    try:
        dt = datetime.fromisoformat(str(ts_str).replace("Z","+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None

def mins_between(t1_str, t2_str):
    """Calculate minutes between two HH:MM:SS strings. Returns None if invalid."""
    try:
        fmt = "%H:%M:%S"
        t1 = datetime.strptime(t1_str[:8], fmt)
        t2 = datetime.strptime(t2_str[:8], fmt)
        diff = (t2 - t1).total_seconds() / 60
        return round(diff, 1) if diff >= 0 else None
    except Exception:
        return None

# ═══════════════════════════════════════════════════════════════
# ENSURE EXCEL FILES EXIST
# ═══════════════════════════════════════════════════════════════
ensure_excel(SCANS_EXCEL, {"scans": SCAN_COLS, "meta": ["key","value"]})
ensure_excel(USERS_EXCEL, {"users": USER_COLS, "meta": ["key","value"]})

# ═══════════════════════════════════════════════════════════════
# IN-MEMORY USER CACHE
# ═══════════════════════════════════════════════════════════════
_userid_to_badge = None

def get_userid_to_badge():
    global _userid_to_badge
    if _userid_to_badge is not None:
        return _userid_to_badge
    df = read_sheet_as_df(USERS_EXCEL, "users")
    if not df.empty and "uid" in df.columns and "badge_no" in df.columns:
        _userid_to_badge = dict(zip(df["uid"], df["badge_no"]))
    else:
        _userid_to_badge = {}
    print(f"User cache loaded: {len(_userid_to_badge)} users")
    return _userid_to_badge

# ═══════════════════════════════════════════════════════════════
# MAIN — called every 30s by server.py
# ═══════════════════════════════════════════════════════════════
def main():
    global _userid_to_badge
    userid_to_badge = get_userid_to_badge()
    fetch_time      = datetime.now(timezone.utc)

    # ── 1. INCREMENTAL USER FETCH ───────────────────────────────
    user_meta       = read_meta(USERS_EXCEL)
    last_user_fetch = parse_ts(user_meta.get("last_user_fetch"))

    if last_user_fetch:
        uq = db.collection(USER_COLLECTION).where(
            filter=firestore.FieldFilter(USER_CREATED_FIELD, ">", last_user_fetch)
        )
    else:
        print("Users: full fetch...")
        uq = db.collection(USER_COLLECTION)

    new_users = []
    for doc in uq.stream():
        d = doc.to_dict()
        created_raw = d.get(USER_CREATED_FIELD)
        created_str = str(to_ist_naive(created_raw)) if created_raw else ""
        new_users.append({
            "uid":          doc.id,
            "badge_no":     str(d.get("badgeNumber","")).strip(),
            "display_name": str(d.get("displayName","")).strip(),
            "created_at":   created_str,
        })

    if new_users:
        # Get existing UIDs to avoid duplicates
        existing_df = read_sheet_as_df(USERS_EXCEL, "users")
        existing_uids = set(existing_df["uid"].tolist()) if not existing_df.empty else set()
        truly_new_users = [u for u in new_users if u["uid"] not in existing_uids]
        if truly_new_users:
            append_rows_to_sheet(USERS_EXCEL, "users", truly_new_users, USER_COLS)
            write_meta(USERS_EXCEL, {**user_meta, "last_user_fetch": fetch_time.isoformat()})
            # Refresh in-memory cache
            for u in truly_new_users:
                _userid_to_badge[u["uid"]] = u["badge_no"]
            userid_to_badge = _userid_to_badge
            print(f"Users: +{len(truly_new_users)} new")
    else:
        if not last_user_fetch:
            write_meta(USERS_EXCEL, {"last_user_fetch": fetch_time.isoformat()})

    # ── 2. INCREMENTAL SCAN FETCH ───────────────────────────────
    scan_meta       = read_meta(SCANS_EXCEL)
    last_scan_fetch = parse_ts(scan_meta.get("last_scan_fetch"))

    if last_scan_fetch:
        sq = db.collection(SCAN_COLLECTION).where(
            filter=firestore.FieldFilter(TIMESTAMP_FIELD, ">", last_scan_fetch)
        )
    else:
        print("Scans: first full fetch...")
        sq = db.collection(SCAN_COLLECTION)

    new_scans = []
    for doc in sq.stream():
        d = doc.to_dict()
        if TIMESTAMP_FIELD not in d:
            continue
        ts_ist     = to_ist_naive(d[TIMESTAMP_FIELD])
        barcode    = str(d.get("barcode","")).strip()
        user_id    = d.get("userId","")
        scanned_by = str(d.get("scannedBy","")).strip()
        if not barcode:
            barcode = userid_to_badge.get(user_id,"")
        sew     = lookup_badge(barcode)
        scanner = lookup_badge(scanned_by)
        new_scans.append({
            "scan_id":           doc.id,
            "badge_no":          barcode,
            "sewadar_name":      sew["name"],
            "sewadar_centre":    sew["centre"],
            "satsang_point":     sew["satsang_point"],
            "department":        sew["department"],
            "gender":            sew["gender"],
            "type":              str(d.get("type","")).strip().upper(),
            "timestamp":         ts_ist.isoformat(),
            "date":              ts_ist.strftime("%Y-%m-%d"),
            "time":              ts_ist.strftime("%H:%M:%S"),
            "scanned_by_badge":  scanned_by,
            "scanned_by_name":   scanner["name"],
            "scanned_by_centre": scanner["centre"],
        })

    print(f"New scans fetched: {len(new_scans)}")

    # Deduplicate + append to Excel
    if new_scans:
        existing_df   = read_sheet_as_df(SCANS_EXCEL, "scans")
        existing_ids  = set(existing_df["scan_id"].tolist()) if not existing_df.empty else set()
        truly_new     = [s for s in new_scans if s["scan_id"] not in existing_ids]
        if truly_new:
            append_rows_to_sheet(SCANS_EXCEL, "scans", truly_new, SCAN_COLS)
            write_meta(SCANS_EXCEL, {**scan_meta, "last_scan_fetch": fetch_time.isoformat()})
            print(f"Appended {len(truly_new)} scans to Excel")
        else:
            print("All fetched scans already in cache.")
            save_last_fetch_only(SCANS_EXCEL, scan_meta, fetch_time)
            return
    else:
        if not last_scan_fetch:
            write_meta(SCANS_EXCEL, {"last_scan_fetch": fetch_time.isoformat()})
        print("No new scans — skipping analysis.")
        return

    # ── 3. LOAD ALL SCANS FROM EXCEL ───────────────────────────
    all_df = read_sheet_as_df(SCANS_EXCEL, "scans")
    if all_df.empty:
        print("No scan data in Excel yet.")
        return

    print(f"Total scans in Excel: {len(all_df)}")

    # ── 4. RUN FULL ANALYSIS ────────────────────────────────────
    analyse_and_write(all_df, fetch_time)


def save_last_fetch_only(path, meta, fetch_time):
    write_meta(path, {**meta, "last_scan_fetch": fetch_time.isoformat()})


def analyse_and_write(all_df, fetch_time):
    """Run complete analysis on all_df and write data.json."""

    # Normalise types
    all_df["timestamp"] = pd.to_datetime(all_df["timestamp"], format="ISO8601", errors="coerce")
    all_df = all_df.dropna(subset=["timestamp"])
    all_df = all_df.sort_values("timestamp").reset_index(drop=True)
    all_df["type"] = all_df["type"].str.upper().str.strip()
    all_df["badge_no"] = all_df["badge_no"].astype(str).str.strip()
    all_df["date"] = all_df["date"].astype(str).str.strip()
    all_df["time"] = all_df["time"].astype(str).str.strip()

    dates = sorted(all_df["date"].unique().tolist())

    # ── Headcount & Gender ──────────────────────────────────────
    headcount_per_day = []
    gender_summary    = []
    for date in dates:
        day_in  = all_df[(all_df["date"]==date) & (all_df["type"]=="IN")]
        day_all = all_df[all_df["date"]==date]
        headcount_per_day.append({
            "date":          date,
            "total_present": int(day_in["badge_no"].nunique()),
            "total_in":      int((day_all["type"]=="IN").sum()),
            "total_out":     int((day_all["type"]=="OUT").sum()),
        })
        gender_summary.append({
            "date":   date,
            "male":   int(day_in[day_in["gender"]=="M"]["badge_no"].nunique()),
            "female": int(day_in[day_in["gender"]=="F"]["badge_no"].nunique()),
        })

    # ── Arrival Slots (30-min buckets per day) ──────────────────
    arrival_slots = {}
    for date in dates:
        day_in = all_df[(all_df["date"]==date) & (all_df["type"]=="IN")]
        slots  = {}
        for _, row in day_in.iterrows():
            try:
                h, m = int(row["time"][:2]), int(row["time"][3:5])
                slot = f"{h:02d}:{(m//30)*30:02d}"
                slots[slot] = slots.get(slot, 0) + 1
            except Exception:
                pass
        arrival_slots[date] = dict(sorted(slots.items()))

    # ── Sewadar Attendance Matrix (FIXED LOGIC) ─────────────────
    # PRESENT = any IN scan that day (even without OUT)
    # ABSENT  = in badge_lookup but zero IN scans today
    sewadar_map = {}
    for _, row in all_df.iterrows():
        badge = row["badge_no"]
        if not badge or badge == "nan":
            continue
        if badge not in sewadar_map:
            sewadar_map[badge] = {
                "badge_no":       badge,
                "sewadar_name":   row["sewadar_name"],
                "sewadar_centre": row["sewadar_centre"],
                "satsang_point":  row.get("satsang_point",""),
                "department":     row["department"],
                "gender":         row["gender"],
            }
        dt = row["date"]
        if dt not in sewadar_map[badge]:
            sewadar_map[badge][dt] = {
                "status":    "PRESENT",   # PRESENT if ANY IN exists — fixed ✅
                "first_in":  None,
                "last_in":   None,
                "last_out":  None,
                "duty_mins": None,
            }
        if row["type"] == "IN":
            cur_first = sewadar_map[badge][dt]["first_in"]
            cur_last  = sewadar_map[badge][dt]["last_in"]
            if not cur_first or row["time"] < cur_first:
                sewadar_map[badge][dt]["first_in"] = row["time"]
            if not cur_last or row["time"] > cur_last:
                sewadar_map[badge][dt]["last_in"] = row["time"]
        elif row["type"] == "OUT":
            cur = sewadar_map[badge][dt]["last_out"]
            if not cur or row["time"] > cur:
                sewadar_map[badge][dt]["last_out"] = row["time"]

    # Compute duty_mins after all scans processed
    for badge, data in sewadar_map.items():
        for dt in dates:
            if dt in data and data[dt].get("first_in") and data[dt].get("last_out"):
                data[dt]["duty_mins"] = mins_between(data[dt]["first_in"], data[dt]["last_out"])

    sewadar_attendance = list(sewadar_map.values())

    # ── Centre Summary (FIXED: uses centre_expected not dept) ───
    centre_map = {}
    for _, row in all_df.iterrows():
        c = row.get("sewadar_centre") or "Unknown"
        if c not in centre_map:
            centre_map[c] = {"centre": c, "dates": {}, "_b": set()}
        dt = row["date"]
        if dt not in centre_map[c]["dates"]:
            centre_map[c]["dates"][dt] = {"_p": set(), "in_count": 0, "out_count": 0}
        if row["type"] == "IN":
            centre_map[c]["dates"][dt]["in_count"] += 1
            centre_map[c]["dates"][dt]["_p"].add(row["badge_no"])
            centre_map[c]["_b"].add(row["badge_no"])
        else:
            centre_map[c]["dates"][dt]["out_count"] += 1

    centre_summary = []
    for c_data in centre_map.values():
        date_out = {}
        for dt, info in c_data["dates"].items():
            date_out[dt] = {
                "present":   len(info["_p"]),
                "in_count":  info["in_count"],
                "out_count": info["out_count"],
            }
        for dt in dates:
            if dt not in date_out:
                date_out[dt] = {"present": 0, "in_count": 0, "out_count": 0}
        centre_summary.append({
            "centre":         c_data["centre"],
            "total_sewadars": len(c_data["_b"]),
            "expected":       centre_expected.get(c_data["centre"], 0),  # FIXED ✅
            "dates":          date_out,
        })

    # ── Absent per Day (FIXED: uses full badge_lookup) ──────────
    # ABSENT = all registered sewadars who have zero IN scans today
    # This catches even sewadars who NEVER scanned at all ✅
    absent_per_day = []
    for date in dates:
        scanned_today = set(
            all_df[(all_df["date"]==date) & (all_df["type"]=="IN")]["badge_no"].tolist()
        )
        absent_list = []
        for b, info in badge_lookup.items():
            if b not in scanned_today:
                absent_list.append({
                    "badge_no":   b,
                    "name":       info["name"],
                    "centre":     info["centre"],
                    "department": info["department"],
                    "gender":     info["gender"],
                    "satsang_point": info["satsang_point"],
                })
        absent_per_day.append({
            "date":         date,
            "absent_count": len(absent_list),
            "absent":       absent_list,
        })

    # ── Department Completion % ─────────────────────────────────
    dept_completion = {}
    for date in dates:
        scanned_today = set(
            all_df[(all_df["date"]==date) & (all_df["type"]=="IN")]["badge_no"].tolist()
        )
        per_dept = {}
        for b, info in badge_lookup.items():
            dept = info.get("department","").strip()
            if not dept:
                continue
            if dept not in per_dept:
                per_dept[dept] = {"expected": 0, "present": 0}
            per_dept[dept]["expected"] += 1
            if b in scanned_today:
                per_dept[dept]["present"] += 1
        dept_completion[date] = {
            d: {
                "expected": v["expected"],
                "present":  v["present"],
                "pct":      round(v["present"]/v["expected"]*100, 1) if v["expected"] else 0,
            }
            for d, v in per_dept.items()
        }

    # ── Centre × Department Pivot Matrix (new report) ──────────
    # Rows: centres, Columns: departments, Values: count of unique badges
    # Built day-wise for the pivot report
    pivot_data = {}
    for date in dates:
        day_in = all_df[(all_df["date"]==date) & (all_df["type"]=="IN")]
        pivot = {}
        for _, row in day_in.iterrows():
            c = row.get("sewadar_centre") or "Unknown"
            d = row.get("department") or "Unknown"
            if c not in pivot:
                pivot[c] = {}
            if d not in pivot[c]:
                pivot[c][d] = set()
            pivot[c][d].add(row["badge_no"])
        # Convert sets to counts
        pivot_data[date] = {
            c: {d: len(badges) for d, badges in depts.items()}
            for c, depts in pivot.items()
        }

    # ── Badge-day grouping for alerts ───────────────────────────
    badge_day_scans = {}
    for _, row in all_df.iterrows():
        key = f"{row['badge_no']}|{row['date']}"
        badge_day_scans.setdefault(key, []).append(row.to_dict())

    duplicate_flags  = []   # IN scanned 2+ times same day
    multi_scan_flags = []   # OUT scanned 2+ times same day (separate from dup IN)
    missing_out      = []   # Has IN but no OUT (STILL PRESENT — just alert)
    orphaned_outs    = []   # Has OUT but no IN
    re_entry_flags   = []   # Complete IN→OUT→IN cycle

    for key, scans_list in badge_day_scans.items():
        badge_no, date = key.split("|", 1)
        ss        = sorted(scans_list, key=lambda s: s["time"])
        in_times  = sorted([s["time"] for s in ss if s["type"]=="IN"])
        out_times = sorted([s["time"] for s in ss if s["type"]=="OUT"])
        name      = ss[0].get("sewadar_name","Unknown")
        centre    = ss[0].get("sewadar_centre","")
        dept      = ss[0].get("department","")

        # FIX L3: duplicate_flags = IN only, multi_scan_flags = OUT only ✅
        if len(in_times) > 1:
            duplicate_flags.append({
                "badge_no":     badge_no, "sewadar_name": name,
                "centre":       centre,   "department":   dept,
                "date":         date,     "in_count":     len(in_times),
                "times":        in_times,
            })
        if len(out_times) > 1:
            multi_scan_flags.append({
                "badge_no":     badge_no, "sewadar_name": name,
                "centre":       centre,   "department":   dept,
                "date":         date,     "type":         "OUT",
                "count":        len(out_times), "times":  out_times,
            })

        has_in  = bool(in_times)
        has_out = bool(out_times)

        # FIX L1: missing_out = PRESENT status, just flagged for alert ✅
        if has_in and not has_out:
            missing_out.append({
                "badge_no":     badge_no, "sewadar_name": name,
                "centre":       centre,   "department":   dept,
                "date":         date,
                "first_in":     in_times[0],
                "last_in":      in_times[-1],   # FIX L5: store last_in too ✅
            })
        elif has_out and not has_in:
            orphaned_outs.append({
                "badge_no":     badge_no, "sewadar_name": name,
                "centre":       centre,   "department":   dept,
                "date":         date,
                "out_times":    out_times,  # FIX L6: store all out times ✅
                "out_time":     out_times[0],
            })

        # FIX L4: Re-entry — detect IN→OUT→IN across any duplicates ✅
        # Walk through time-sorted scans, track last seen type
        last_type     = None
        had_complete_cycle = False
        cycle_first_in = None
        cycle_out_time = None
        for sc in ss:
            t = sc["type"]
            if t == "IN" and last_type is None:
                cycle_first_in = sc["time"]
                last_type = "IN"
            elif t == "OUT" and last_type == "IN":
                cycle_out_time = sc["time"]
                last_type = "OUT"
            elif t == "IN" and last_type == "OUT":
                # Complete IN→OUT→IN cycle found
                re_entry_flags.append({
                    "badge_no":     badge_no, "sewadar_name": name,
                    "centre":       centre,   "department":   dept,
                    "date":         date,
                    "first_in":     cycle_first_in,
                    "out_time":     cycle_out_time,
                    "second_in":    sc["time"],
                })
                had_complete_cycle = True
                break

    # ── Scanner Performance ─────────────────────────────────────
    scanner_stats = {}
    for _, row in all_df.iterrows():
        sb = row.get("scanned_by_badge","")
        if not sb or sb == "nan":
            continue
        if sb not in scanner_stats:
            scanner_stats[sb] = {
                "badge":          sb,
                "name":           row.get("scanned_by_name",""),
                "centre":         row.get("scanned_by_centre",""),
                "total_scans":    0,
                "badges_scanned": set(),
                "by_date":        {},
            }
        scanner_stats[sb]["total_scans"] += 1
        scanner_stats[sb]["badges_scanned"].add(row["badge_no"])
        dt = row["date"]
        scanner_stats[sb]["by_date"][dt] = scanner_stats[sb]["by_date"].get(dt, 0) + 1

    scanner_performance = [
        {
            "badge":           v["badge"],
            "name":            v["name"],
            "centre":          v["centre"],
            "total_scans":     v["total_scans"],
            "unique_sewadars": len(v["badges_scanned"]),
            "by_date":         v["by_date"],
        }
        for v in sorted(scanner_stats.values(), key=lambda x: -x["total_scans"])
    ]

    # ── Summary Totals ──────────────────────────────────────────
    total_sewadars = int(all_df["badge_no"].nunique())
    last_scan_ts   = all_df["timestamp"].max()
    last_scan_str  = str(last_scan_ts) if pd.notna(last_scan_ts) else None

    # ── Build final output ──────────────────────────────────────
    output = clean({
        "last_updated":        fetch_time.strftime("%d-%m-%Y %I:%M %p IST"),
        "last_scan_at":        last_scan_str,
        "collection":          SCAN_COLLECTION,
        "total_scans":         len(all_df),
        "total_sewadars":      total_sewadars,
        "total_registered":    len(badge_lookup),
        "department_expected": department_expected,
        "centre_expected":     centre_expected,
        "dates":               dates,
        "headcount_per_day":   headcount_per_day,
        "gender_summary":      gender_summary,
        "arrival_slots":       arrival_slots,
        "centre_summary":      centre_summary,
        "sewadar_attendance":  sewadar_attendance,
        "absent_per_day":      absent_per_day,
        "dept_completion":     dept_completion,
        "pivot_data":          pivot_data,
        "scanner_performance": scanner_performance,
        "duplicate_flags":     duplicate_flags,
        "multi_scan_flags":    multi_scan_flags,
        "missing_out":         missing_out,
        "orphaned_outs":       orphaned_outs,
        "re_entry_flags":      re_entry_flags,
        "scans":               all_df.to_dict("records"),
    })

    # Atomic write
    tmp = OUTPUT_JSON + ".tmp"
    with open(tmp, "w") as f:
        json.dump(output, f, separators=(",",":"), default=str)  # compact = smaller file
    os.replace(tmp, OUTPUT_JSON)

    print(
        f"data.json written — {len(all_df)} scans | "
        f"{len(duplicate_flags)} dup-INs | {len(missing_out)} missing-OUTs | "
        f"{len(re_entry_flags)} re-entries | "
        f"{sum(d['absent_count'] for d in absent_per_day)} absent-slots"
    )


def cold_start_render():
    """Called on startup — load Excel and write data.json immediately (0 Firestore reads)."""
    print("Cold start: loading from Excel...")
    df = read_sheet_as_df(SCANS_EXCEL, "scans")
    if df.empty:
        print("No cached scans yet — will populate on first fetch.")
        return
    analyse_and_write(df, datetime.now(timezone.utc))
    print("Cold start render complete.")


if __name__ == "__main__":
    cold_start_render()
    main()
